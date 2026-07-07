from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
from shutil import copy2
from typing import Any
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from config.excel_mapping import (
    EXCEL_MAPPING,
    get_expected_workbook_path,
    normalize_supplier,
    resolve_month_sheet_name,
    normalize_excel_text
)
from settlement_automation.services.reconciliation import (
    summarize_mobile_adjustments,
    summarize_valero_pay_plus_adjustments,
)

from settlement_automation.models import (
    DailySettlementTotal,
    MobileAdjustment,
    ParsedReport,
    ValeroPayPlusAdjustment,
    ValeroMonthlyCharge,
)

@dataclass(frozen=True)
class ExcelFeeValidation:
    supplier: str
    location_id: str
    location_name: str
    business_date: date
    source: str
    gross_amt: Decimal
    net_amt: Decimal
    parsed_fees: Decimal
    calculated_fees: Decimal
    difference: Decimal
    is_valid: bool

@dataclass(frozen=True)
class ExcelTarget:
    supplier: str
    location_id: str
    location_name: str
    business_date: date
    workbook_path: Path
    sheet_name: str | None


@dataclass(frozen=True)
class ExcelPlannedValue:
    target: ExcelTarget
    field_name: str
    column_header: str
    value: Decimal
    source: str
    mode: str


@dataclass
class ExcelWritePlan:
    report_supplier: str
    report_date: date
    planned_values: list[ExcelPlannedValue] = field(default_factory=list)
    fee_validations: list[ExcelFeeValidation] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def has_warnings(self) -> bool:
        return bool(self.warnings)

    @property
    def affected_workbooks(self) -> set[Path]:
        return {planned.target.workbook_path for planned in self.planned_values}

@dataclass
class ExcelWriteResult:
    dry_run: bool
    write_originals: bool
    plan: ExcelWritePlan
    resolution: ExcelPlanResolution | None = None
    apply_result: ExcelApplyResult | None = None
    written_count: int = 0
    skipped_count: int = 0
    warnings: list[str] = field(default_factory=list)

@dataclass(frozen=True)
class ExcelResolvedValue:
    planned_value: ExcelPlannedValue
    workbook_path: Path
    sheet_name: str
    header_row: int
    row_number: int
    column_number: int
    cell_ref: str
    current_value: object


@dataclass(frozen=True)
class ExcelFeeFormulaValidation:
    supplier: str
    location_id: str
    location_name: str
    business_date: date
    workbook_path: Path
    sheet_name: str
    cell_ref: str
    current_value: object
    is_formula: bool


@dataclass
class ExcelPlanResolution:
    resolved_values: list[ExcelResolvedValue] = field(default_factory=list)
    fee_formula_validations: list[ExcelFeeFormulaValidation] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return not self.warnings


@dataclass(frozen=True)
class ExcelAppliedChange:
    supplier: str
    location_id: str
    location_name: str
    business_date: date
    workbook_path: Path
    output_path: Path
    sheet_name: str
    cell_ref: str
    field_name: str
    column_header: str
    source: str
    mode: str
    old_value: object
    new_value: object
    status: str


@dataclass
class ExcelApplyResult:
    changes: list[ExcelAppliedChange] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def written_count(self) -> int:
        return sum(1 for change in self.changes if change.status == "written")

    @property
    def skipped_count(self) -> int:
        return sum(1 for change in self.changes if change.status.startswith("skipped"))



def _get_primary_daily_total_dates(report: ParsedReport) -> set[date]:
    """
    For normal supplier reports, the latest daily total date in the report
    is the current full-day settlement date.

    Older daily_total rows in the same report are late/backdated adjustments
    and should be added to the existing workbook value, not replace it.

    Valero is handled separately because it already separates mobile/pay+/monthly
    adjustment buckets and can have multi-day business dates in parser logic.
    """
    if not report.daily_totals:
        return set()

    supplier = normalize_supplier(report.supplier)

    if supplier == "VALERO":
        return {row.date for row in report.daily_totals}

    latest_date = max(row.date for row in report.daily_totals)
    return {latest_date}





def _is_excel_temp_file(path: Path) -> bool:
    return path.name.startswith("~$")


def _summarize_valero_monthly_charges(
    rows: list[ValeroMonthlyCharge],
) -> list[ValeroMonthlyCharge]:
    grouped: dict[tuple[str, str, str, date], Decimal] = defaultdict(
        lambda: Decimal("0.00")
    )
    descriptions: dict[tuple[str, str, str, date], list[str]] = defaultdict(list)

    for row in rows:
        key = (row.supplier, row.location_id, row.location_name, row.date)
        grouped[key] += row.amount
        descriptions[key].append(row.description)

    summary: list[ValeroMonthlyCharge] = []

    for key, amount in grouped.items():
        supplier, location_id, location_name, charge_date = key

        summary.append(
            ValeroMonthlyCharge(
                supplier=supplier,
                location_id=location_id,
                location_name=location_name,
                date=charge_date,
                amount=amount,
                description="; ".join(sorted(set(descriptions[key]))),
            )
        )

    return sorted(summary, key=lambda row: (row.date, row.location_id))


def _monthly_charge_group_key(
    resolved: ExcelResolvedValue,
) -> tuple[Path, str, int, str, str, date]:
    planned = resolved.planned_value
    target = planned.target

    return (
        resolved.workbook_path,
        resolved.sheet_name,
        resolved.row_number,
        target.supplier,
        target.location_id,
        target.business_date,
    )

def _append_valero_monthly_charge_values(
    *,
    plan: ExcelWritePlan,
    workbook_root: Path,
    row: ValeroMonthlyCharge,
) -> None:
    columns = EXCEL_MAPPING.columns

    target = _build_target_without_opening_workbook(
        workbook_root=workbook_root,
        supplier=row.supplier,
        location_id=row.location_id,
        location_name=row.location_name,
        business_date=row.date,
    )

    plan.planned_values.extend(
        [
            ExcelPlannedValue(
                target=target,
                field_name="monthly_valero_charges",
                column_header=columns.monthly_valero_charges,
                value=row.amount,
                source="valero_monthly_charge_summary",
                mode="set",
            ),
            ExcelPlannedValue(
                target=target,
                field_name="fees",
                column_header=columns.fees,
                value=row.amount,
                source="valero_monthly_charge_summary",
                mode="add_monthly_charge_to_fee_formula",
            ),
        ]
    )

def _apply_valero_monthly_charge_group(
    *,
    wb,
    output_path: Path,
    group_values: list[ExcelResolvedValue],
    warnings: list[str],
) -> list[ExcelAppliedChange]:
    changes: list[ExcelAppliedChange] = []

    by_field = {
        resolved.planned_value.field_name: resolved
        for resolved in group_values
    }

    monthly_resolved = by_field.get("monthly_valero_charges")
    fee_resolved = by_field.get("fees")

    if monthly_resolved is None or fee_resolved is None:
        warning = (
            f"Incomplete monthly charge group in workbook={output_path.name}: "
            f"fields={sorted(by_field)}"
        )
        warnings.append(warning)
        return changes

    ws = wb[monthly_resolved.sheet_name]

    monthly_cell = ws[monthly_resolved.cell_ref]
    fee_cell = ws[fee_resolved.cell_ref]

    monthly_old = monthly_cell.value
    fee_old = fee_cell.value

    planned_amount = monthly_resolved.planned_value.value
    monthly_new = _to_excel_number(planned_amount)

    existing_monthly = _coerce_decimal(monthly_old)
    monthly_already_set = _decimal_equal(existing_monthly, planned_amount)

    target = monthly_resolved.planned_value.target

    # Always make the monthly column deterministic/idempotent.
    monthly_cell.value = monthly_new

    monthly_planned = monthly_resolved.planned_value
    changes.append(
        ExcelAppliedChange(
            supplier=target.supplier,
            location_id=target.location_id,
            location_name=target.location_name,
            business_date=target.business_date,
            workbook_path=monthly_resolved.workbook_path,
            output_path=output_path,
            sheet_name=monthly_resolved.sheet_name,
            cell_ref=monthly_resolved.cell_ref,
            field_name=monthly_planned.field_name,
            column_header=monthly_planned.column_header,
            source=monthly_planned.source,
            mode=monthly_planned.mode,
            old_value=monthly_old,
            new_value=monthly_new,
            status="written",
        )
    )

    fee_planned = fee_resolved.planned_value

    if _is_formula_value(fee_old):
        if _formula_references_cell(fee_old, monthly_resolved.cell_ref):
            fee_new = fee_old
            status = "skipped_formula_already_references_monthly_charge"
        else:
            fee_new = _append_cell_to_formula(str(fee_old), monthly_resolved.cell_ref)
            fee_cell.value = fee_new
            status = "written"

        changes.append(
            ExcelAppliedChange(
                supplier=target.supplier,
                location_id=target.location_id,
                location_name=target.location_name,
                business_date=target.business_date,
                workbook_path=fee_resolved.workbook_path,
                output_path=output_path,
                sheet_name=fee_resolved.sheet_name,
                cell_ref=fee_resolved.cell_ref,
                field_name=fee_planned.field_name,
                column_header=fee_planned.column_header,
                source=fee_planned.source,
                mode=fee_planned.mode,
                old_value=fee_old,
                new_value=fee_new,
                status=status,
            )
        )

        return changes

    existing_fee = _coerce_decimal(fee_old)

    if existing_fee is None:
        warning = (
            f"Cannot add monthly charge to non-numeric/non-formula CC Fee cell: "
            f"workbook={output_path.name}, sheet={fee_resolved.sheet_name}, "
            f"date={target.business_date}, location={target.location_id}, "
            f"cell={fee_resolved.cell_ref}, value={fee_old!r}"
        )
        warnings.append(warning)

        changes.append(
            ExcelAppliedChange(
                supplier=target.supplier,
                location_id=target.location_id,
                location_name=target.location_name,
                business_date=target.business_date,
                workbook_path=fee_resolved.workbook_path,
                output_path=output_path,
                sheet_name=fee_resolved.sheet_name,
                cell_ref=fee_resolved.cell_ref,
                field_name=fee_planned.field_name,
                column_header=fee_planned.column_header,
                source=fee_planned.source,
                mode=fee_planned.mode,
                old_value=fee_old,
                new_value=fee_old,
                status="skipped_non_numeric_fee",
            )
        )

        return changes

    # Numeric fallback:
    # If monthly column already had this exact amount, assume previous run already
    # added it to CC Fee. Do not double-add.
    if monthly_already_set:
        fee_new = fee_old
        status = "skipped_already_applied"
    else:
        fee_new = _to_excel_number(existing_fee + planned_amount)
        fee_cell.value = fee_new
        status = "written"

    changes.append(
        ExcelAppliedChange(
            supplier=target.supplier,
            location_id=target.location_id,
            location_name=target.location_name,
            business_date=target.business_date,
            workbook_path=fee_resolved.workbook_path,
            output_path=output_path,
            sheet_name=fee_resolved.sheet_name,
            cell_ref=fee_resolved.cell_ref,
            field_name=fee_planned.field_name,
            column_header=fee_planned.column_header,
            source=fee_planned.source,
            mode=fee_planned.mode,
            old_value=fee_old,
            new_value=fee_new,
            status=status,
        )
    )

    return changes


def _apply_valero_pay_plus_group(
    *,
    wb,
    output_path: Path,
    group_values: list[ExcelResolvedValue],
    warnings: list[str],
) -> list[ExcelAppliedChange]:
    changes: list[ExcelAppliedChange] = []

    by_field = {
        resolved.planned_value.field_name: resolved
        for resolved in group_values
    }

    gross_resolved = by_field.get("gross_amt")
    net_resolved = by_field.get("net_amt")
    pay_plus_resolved = by_field.get("valero_pay_plus")

    if gross_resolved is None or net_resolved is None or pay_plus_resolved is None:
        warning = (
            f"Incomplete Valero Pay+ group in workbook={output_path.name}: "
            f"fields={sorted(by_field)}"
        )
        warnings.append(warning)
        return changes

    ws = wb[gross_resolved.sheet_name]

    gross_cell = ws[gross_resolved.cell_ref]
    net_cell = ws[net_resolved.cell_ref]
    pay_plus_cell = ws[pay_plus_resolved.cell_ref]

    gross_old = gross_cell.value
    net_old = net_cell.value
    pay_plus_old = pay_plus_cell.value

    planned_amount = pay_plus_resolved.planned_value.value
    target = gross_resolved.planned_value.target

    existing_pay_plus = _coerce_decimal(pay_plus_old)

    already_applied = _decimal_equal(existing_pay_plus, planned_amount)
    has_different_existing_pay_plus = (
        existing_pay_plus is not None
        and not _is_effectively_zero(existing_pay_plus)
        and not already_applied
    )

    if already_applied:
        pay_plus_cell.value = _to_excel_number(planned_amount)

        for resolved in [gross_resolved, net_resolved, pay_plus_resolved]:
            planned = resolved.planned_value
            cell = ws[resolved.cell_ref]

            changes.append(
                ExcelAppliedChange(
                    supplier=target.supplier,
                    location_id=target.location_id,
                    location_name=target.location_name,
                    business_date=target.business_date,
                    workbook_path=resolved.workbook_path,
                    output_path=output_path,
                    sheet_name=resolved.sheet_name,
                    cell_ref=resolved.cell_ref,
                    field_name=planned.field_name,
                    column_header=planned.column_header,
                    source=planned.source,
                    mode=planned.mode,
                    old_value=cell.value,
                    new_value=cell.value,
                    status="skipped_already_applied",
                )
            )

        return changes

    if has_different_existing_pay_plus:
        warning = (
            f"Existing Valero Pay+ value differs from planned value. "
            f"Skipping Pay+ add to avoid double-counting/corruption: "
            f"workbook={output_path.name}, sheet={gross_resolved.sheet_name}, "
            f"date={target.business_date}, location={target.location_id}, "
            f"existing_pay_plus={existing_pay_plus}, planned_pay_plus={planned_amount}"
        )
        warnings.append(warning)

        for resolved in [gross_resolved, net_resolved, pay_plus_resolved]:
            planned = resolved.planned_value
            cell = ws[resolved.cell_ref]

            changes.append(
                ExcelAppliedChange(
                    supplier=target.supplier,
                    location_id=target.location_id,
                    location_name=target.location_name,
                    business_date=target.business_date,
                    workbook_path=resolved.workbook_path,
                    output_path=output_path,
                    sheet_name=resolved.sheet_name,
                    cell_ref=resolved.cell_ref,
                    field_name=planned.field_name,
                    column_header=planned.column_header,
                    source=planned.source,
                    mode=planned.mode,
                    old_value=cell.value,
                    new_value=cell.value,
                    status="skipped_existing_pay_plus_mismatch",
                )
            )

        return changes

    gross_existing = _coerce_decimal(gross_old)
    net_existing = _coerce_decimal(net_old)

    if gross_existing is None or net_existing is None:
        warning = (
            f"Cannot apply Valero Pay+ because Gross/NET cell is not numeric: "
            f"workbook={output_path.name}, sheet={gross_resolved.sheet_name}, "
            f"date={target.business_date}, location={target.location_id}, "
            f"gross={gross_old!r}, net={net_old!r}"
        )
        warnings.append(warning)
        return changes

    if _is_formula_value(gross_old) or _is_formula_value(net_old):
        warning = (
            f"Cannot apply Valero Pay+ over formula Gross/NET cells: "
            f"workbook={output_path.name}, sheet={gross_resolved.sheet_name}, "
            f"date={target.business_date}, location={target.location_id}, "
            f"gross={gross_old!r}, net={net_old!r}"
        )
        warnings.append(warning)
        return changes

    gross_new = _to_excel_number(gross_existing + planned_amount)
    net_new = _to_excel_number(net_existing + planned_amount)
    pay_plus_new = _to_excel_number(planned_amount)

    gross_cell.value = gross_new
    net_cell.value = net_new
    pay_plus_cell.value = pay_plus_new

    for resolved, old_value, new_value in [
        (gross_resolved, gross_old, gross_new),
        (net_resolved, net_old, net_new),
        (pay_plus_resolved, pay_plus_old, pay_plus_new),
    ]:
        planned = resolved.planned_value

        changes.append(
            ExcelAppliedChange(
                supplier=target.supplier,
                location_id=target.location_id,
                location_name=target.location_name,
                business_date=target.business_date,
                workbook_path=resolved.workbook_path,
                output_path=output_path,
                sheet_name=resolved.sheet_name,
                cell_ref=resolved.cell_ref,
                field_name=planned.field_name,
                column_header=planned.column_header,
                source=planned.source,
                mode=planned.mode,
                old_value=old_value,
                new_value=new_value,
                status="written",
            )
        )

    return changes



def _resolve_existing_workbook_path(expected_path: Path) -> tuple[Path | None, list[str]]:
    warnings: list[str] = []

    if expected_path.exists() and not _is_excel_temp_file(expected_path):
        return expected_path, warnings

    workbook_dir = expected_path.parent
    if not workbook_dir.exists():
        warnings.append(f"Workbook directory does not exist: {workbook_dir}")
        return None, warnings

    expected_normalized = normalize_excel_text(expected_path.stem)

    candidates = [
        path
        for path in workbook_dir.glob("*.xlsx")
        if not _is_excel_temp_file(path)
    ]

    normalized_matches = [
        path
        for path in candidates
        if normalize_excel_text(path.stem) == expected_normalized
    ]

    if len(normalized_matches) == 1:
        warnings.append(
            f"Workbook exact path missing, but normalized match found: "
            f"expected={expected_path.name}, matched={normalized_matches[0].name}"
        )
        return normalized_matches[0], warnings

    scored_candidates = []
    for path in candidates:
        score = SequenceMatcher(
            None,
            expected_normalized,
            normalize_excel_text(path.stem),
        ).ratio()
        scored_candidates.append((score, path))

    scored_candidates.sort(reverse=True, key=lambda item: item[0])

    if scored_candidates and scored_candidates[0][0] >= 0.92:
        score, matched_path = scored_candidates[0]
        warnings.append(
            f"Workbook fuzzy match used: expected={expected_path.name}, "
            f"matched={matched_path.name}, score={score:.3f}. "
            f"Consider adding an override in config/excel_mapping.py."
        )
        return matched_path, warnings

    warnings.append(f"Workbook not found: {expected_path}")
    return None, warnings


def _coerce_excel_date(value: object) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, int | float):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%d-%B-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue

    return None


def _find_header_row_and_columns(
    ws,
    required_headers: set[str],
    *,
    max_rows: int,
) -> tuple[int, dict[str, int]] | None:
    normalized_required = {
        normalize_excel_text(header)
        for header in required_headers
    }

    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        header_map: dict[str, int] = {}

        for cell in ws[row_idx]:
            normalized_value = normalize_excel_text(cell.value)
            if normalized_value:
                header_map[normalized_value] = cell.column

        if normalized_required.issubset(set(header_map)):
            return row_idx, header_map

    return None


# def _find_date_row(
#     ws,
#     *,
#     date_column_number: int,
#     header_row: int,
#     target_date: date,
# ) -> int | None:
#     for row_idx in range(header_row + 1, ws.max_row + 1):
#         cell_value = ws.cell(row=row_idx, column=date_column_number).value
#         if _coerce_excel_date(cell_value) == target_date:
#             return row_idx
#
#     return None

def _find_date_row(
    ws_values,
    *,
    date_column_number: int,
    header_row: int,
    target_date: date,
) -> int | None:
    """
    Find date row using a data_only worksheet.

    Many office workbooks have Date cells like:
        A4 = A3 + 1

    With data_only=False, openpyxl sees the formula string.
    With data_only=True, openpyxl sees the cached calculated date.
    """
    for row_idx in range(header_row + 1, ws_values.max_row + 1):
        cell_value = ws_values.cell(row=row_idx, column=date_column_number).value

        if _coerce_excel_date(cell_value) == target_date:
            return row_idx

    return None

def _is_formula_value(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")

def _cell_ref_variants(cell_ref: str) -> set[str]:
    """
    Example:
        H12 -> {"H12", "$H12", "H$12", "$H$12"}
    """
    match = re.match(r"^([A-Z]+)(\d+)$", cell_ref.upper())

    if not match:
        return {cell_ref.upper()}

    col, row = match.groups()

    return {
        f"{col}{row}",
        f"${col}{row}",
        f"{col}${row}",
        f"${col}${row}",
    }


def _formula_references_cell(formula: object, cell_ref: str) -> bool:
    if not _is_formula_value(formula):
        return False

    normalized_formula = str(formula).upper().replace(" ", "")

    return any(
        variant in normalized_formula
        for variant in _cell_ref_variants(cell_ref)
    )


def _append_cell_to_formula(formula: str, cell_ref: str) -> str:
    formula_body = formula[1:] if formula.startswith("=") else formula
    return f"=({formula_body})+{cell_ref}"

def _build_target_without_opening_workbook(
    *,
    workbook_root: Path,
    supplier: str,
    location_id: str,
    location_name: str,
    business_date: date,
) -> ExcelTarget:
    supplier = normalize_supplier(supplier)
    workbook_path = get_expected_workbook_path(
        workbook_root=workbook_root,
        supplier=supplier,
        location_id=location_id,
        location_name=location_name,
        year=business_date.year,
    )

    # Sheet name cannot be confirmed until the workbook is opened.
    # For now, use the first expected candidate as a placeholder.
    sheet_name = None

    return ExcelTarget(
        supplier=supplier,
        location_id=str(location_id),
        location_name=location_name,
        business_date=business_date,
        workbook_path=workbook_path,
        sheet_name=sheet_name,
    )


def _append_daily_total_values(
    *,
    plan: ExcelWritePlan,
    workbook_root: Path,
    row: DailySettlementTotal,
    mode: str = "set",
    source: str = "daily_total",
) -> None:
    columns = EXCEL_MAPPING.columns

    _validate_fee_math(
        plan=plan,
        supplier=row.supplier,
        location_id=row.location_id,
        location_name=row.location_name,
        business_date=row.date,
        gross_amt=row.gross_amt,
        net_amt=row.net_amt,
        parsed_fees=row.fees,
        source=source,
    )

    target = _build_target_without_opening_workbook(
        workbook_root=workbook_root,
        supplier=row.supplier,
        location_id=row.location_id,
        location_name=row.location_name,
        business_date=row.date,
    )

    plan.planned_values.extend(
        [
            ExcelPlannedValue(
                target=target,
                field_name="gross_amt",
                column_header=columns.gross_amt,
                value=row.gross_amt,
                source=source,
                mode=mode,
            ),
            ExcelPlannedValue(
                target=target,
                field_name="net_amt",
                column_header=columns.net_amt,
                value=row.net_amt,
                source=source,
                mode=mode,
            ),
        ]
    )


def _append_valero_mobile_summary_values(
    *,
    plan: ExcelWritePlan,
    workbook_root: Path,
    row: MobileAdjustment,
) -> None:
    columns = EXCEL_MAPPING.columns

    _validate_fee_math(
        plan=plan,
        supplier=row.supplier,
        location_id=row.location_id,
        location_name=row.location_name,
        business_date=row.date,
        gross_amt=row.gross_amt,
        net_amt=row.net_amt,
        parsed_fees=row.fees,
        source="mobile_adjustment_summary",
    )

    target = _build_target_without_opening_workbook(
        workbook_root=workbook_root,
        supplier=row.supplier,
        location_id=row.location_id,
        location_name=row.location_name,
        business_date=row.date,
    )

    # Do not write CC Fee.
    # The workbook formula calculates it automatically.
    plan.planned_values.extend(
        [
            ExcelPlannedValue(
                target=target,
                field_name="gross_amt",
                column_header=columns.gross_amt,
                value=row.gross_amt,
                source="mobile_adjustment_summary",
                mode="add_to_base",
            ),
            ExcelPlannedValue(
                target=target,
                field_name="net_amt",
                column_header=columns.net_amt,
                value=row.net_amt,
                source="mobile_adjustment_summary",
                mode="add_to_base",
            ),
            ExcelPlannedValue(
                target=target,
                field_name="mobile_pay_added_to_gross_net",
                column_header=columns.mobile_pay,
                value=row.net_amt,
                source="mobile_adjustment_summary",
                mode="set",
            ),
        ]
    )

def _append_valero_pay_plus_values(
    *,
    plan: ExcelWritePlan,
    workbook_root: Path,
    row: ValeroPayPlusAdjustment,
) -> None:
    columns = EXCEL_MAPPING.columns

    target = _build_target_without_opening_workbook(
        workbook_root=workbook_root,
        supplier=row.supplier,
        location_id=row.location_id,
        location_name=row.location_name,
        business_date=row.date,
    )

    plan.planned_values.extend(
        [
            ExcelPlannedValue(
                target=target,
                field_name="gross_amt",
                column_header=columns.gross_amt,
                value=row.amount,
                source="valero_pay_plus_summary",
                mode="add_to_base",
            ),
            ExcelPlannedValue(
                target=target,
                field_name="net_amt",
                column_header=columns.net_amt,
                value=row.amount,
                source="valero_pay_plus_summary",
                mode="add_to_base",
            ),
            ExcelPlannedValue(
                target=target,
                field_name="valero_pay_plus",
                column_header=columns.valero_pay_plus,
                value=row.amount,
                source="valero_pay_plus_summary",
                mode="set",
            ),
        ]
    )



def _to_excel_number(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01")))


def _coerce_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, int | float):
        return Decimal(str(value))

    if isinstance(value, str):
        text = value.strip()

        if text.startswith("="):
            return None

        text = text.replace(",", "").replace("$", "")

        if text == "":
            return None

        try:
            return Decimal(text)
        except Exception:
            return None

    return None


def _is_effectively_zero(value: Decimal | None) -> bool:
    return value is None or abs(value) <= Decimal("0.005")


def _decimal_equal(a: Decimal | None, b: Decimal, tolerance: Decimal = Decimal("0.01")) -> bool:
    if a is None:
        return False
    return abs(a - b) <= tolerance


def _format_change_value(value: object) -> object:
    if isinstance(value, Decimal):
        return _to_excel_number(value)
    return value

def _get_output_workbook_path(
    *,
    original_workbook_path: Path,
    output_root: Path,
    report_date: date,
) -> Path:
    return output_root / report_date.isoformat() / original_workbook_path.name


def _copy_workbooks_to_output(
    *,
    workbook_paths: set[Path],
    output_root: Path,
    report_date: date,
    overwrite: bool = True,
) -> dict[Path, Path]:
    output_paths: dict[Path, Path] = {}

    for workbook_path in workbook_paths:
        output_path = _get_output_workbook_path(
            original_workbook_path=workbook_path,
            output_root=output_root,
            report_date=report_date,
        )

        output_path.parent.mkdir(parents=True, exist_ok=True)

        if output_path.exists() and not overwrite:
            output_paths[workbook_path] = output_path
            continue

        copy2(workbook_path, output_path)
        output_paths[workbook_path] = output_path

    return output_paths

AUTOMATION_LOG_SHEET_NAME = "_SETTLEMENT_AUTOMATION_LOG"


def _get_or_create_automation_log_sheet(wb):
    if AUTOMATION_LOG_SHEET_NAME in wb.sheetnames:
        ws = wb[AUTOMATION_LOG_SHEET_NAME]
    else:
        ws = wb.create_sheet(AUTOMATION_LOG_SHEET_NAME)
        ws.append(
            [
                "operation_id",
                "supplier",
                "location_id",
                "business_date",
                "field_name",
                "source",
                "amount",
            ]
        )
        ws.sheet_state = "hidden"

    return ws


def _build_operation_id(resolved: ExcelResolvedValue) -> str:
    planned = resolved.planned_value
    target = planned.target

    return "|".join(
        [
            target.supplier,
            target.location_id,
            target.business_date.isoformat(),
            planned.field_name,
            planned.source,
            str(planned.value),
        ]
    )


def _operation_already_logged(wb, operation_id: str) -> bool:
    if AUTOMATION_LOG_SHEET_NAME not in wb.sheetnames:
        return False

    ws = wb[AUTOMATION_LOG_SHEET_NAME]

    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and row[0] == operation_id:
            return True

    return False


def _log_operation(wb, resolved: ExcelResolvedValue) -> None:
    planned = resolved.planned_value
    target = planned.target
    ws = _get_or_create_automation_log_sheet(wb)

    ws.append(
        [
            _build_operation_id(resolved),
            target.supplier,
            target.location_id,
            target.business_date.isoformat(),
            planned.field_name,
            planned.source,
            str(planned.value),
        ]
    )


def _apply_backdated_daily_total(
    *,
    wb,
    output_path: Path,
    resolved: ExcelResolvedValue,
    warnings: list[str],
) -> ExcelAppliedChange:
    planned = resolved.planned_value
    target = planned.target
    ws = wb[resolved.sheet_name]
    cell = ws[resolved.cell_ref]
    old_value = cell.value

    operation_id = _build_operation_id(resolved)

    if _operation_already_logged(wb, operation_id):
        return ExcelAppliedChange(
            supplier=target.supplier,
            location_id=target.location_id,
            location_name=target.location_name,
            business_date=target.business_date,
            workbook_path=resolved.workbook_path,
            output_path=output_path,
            sheet_name=resolved.sheet_name,
            cell_ref=resolved.cell_ref,
            field_name=planned.field_name,
            column_header=planned.column_header,
            source=planned.source,
            mode=planned.mode,
            old_value=old_value,
            new_value=old_value,
            status="skipped_already_logged",
        )

    existing_value = _coerce_decimal(old_value)

    if existing_value is None:
        # If the older row is blank, use the adjustment amount as the value.
        # This can happen if reports are processed out of chronological order.
        new_value = _to_excel_number(planned.value)
    else:
        new_value = _to_excel_number(existing_value + planned.value)

    if _is_formula_value(old_value):
        warning = (
            f"Cannot apply backdated daily total over formula cell: "
            f"workbook={output_path.name}, sheet={resolved.sheet_name}, "
            f"date={target.business_date}, location={target.location_id}, "
            f"cell={resolved.cell_ref}, value={old_value!r}"
        )
        warnings.append(warning)

        return ExcelAppliedChange(
            supplier=target.supplier,
            location_id=target.location_id,
            location_name=target.location_name,
            business_date=target.business_date,
            workbook_path=resolved.workbook_path,
            output_path=output_path,
            sheet_name=resolved.sheet_name,
            cell_ref=resolved.cell_ref,
            field_name=planned.field_name,
            column_header=planned.column_header,
            source=planned.source,
            mode=planned.mode,
            old_value=old_value,
            new_value=old_value,
            status="skipped_formula",
        )

    cell.value = new_value
    _log_operation(wb, resolved)

    return ExcelAppliedChange(
        supplier=target.supplier,
        location_id=target.location_id,
        location_name=target.location_name,
        business_date=target.business_date,
        workbook_path=resolved.workbook_path,
        output_path=output_path,
        sheet_name=resolved.sheet_name,
        cell_ref=resolved.cell_ref,
        field_name=planned.field_name,
        column_header=planned.column_header,
        source=planned.source,
        mode=planned.mode,
        old_value=old_value,
        new_value=new_value,
        status="written",
    )


def _apply_set_value(
    *,
    ws,
    resolved: ExcelResolvedValue,
    output_path: Path,
    warnings: list[str],
) -> ExcelAppliedChange:
    planned = resolved.planned_value
    target = planned.target

    cell = ws[resolved.cell_ref]
    old_value = cell.value

    if _is_formula_value(old_value) and not EXCEL_MAPPING.writer_policy.overwrite_formulas:
        warning = (
            f"Skipped formula cell write: workbook={output_path.name}, "
            f"sheet={resolved.sheet_name}, cell={resolved.cell_ref}, "
            f"field={planned.field_name}, value={old_value!r}"
        )
        warnings.append(warning)

        return ExcelAppliedChange(
            supplier=target.supplier,
            location_id=target.location_id,
            location_name=target.location_name,
            business_date=target.business_date,
            workbook_path=resolved.workbook_path,
            output_path=output_path,
            sheet_name=resolved.sheet_name,
            cell_ref=resolved.cell_ref,
            field_name=planned.field_name,
            column_header=planned.column_header,
            source=planned.source,
            mode=planned.mode,
            old_value=old_value,
            new_value=old_value,
            status="skipped_formula",
        )

    new_value = _to_excel_number(planned.value)
    cell.value = new_value

    return ExcelAppliedChange(
        supplier=target.supplier,
        location_id=target.location_id,
        location_name=target.location_name,
        business_date=target.business_date,
        workbook_path=resolved.workbook_path,
        output_path=output_path,
        sheet_name=resolved.sheet_name,
        cell_ref=resolved.cell_ref,
        field_name=planned.field_name,
        column_header=planned.column_header,
        source=planned.source,
        mode=planned.mode,
        old_value=old_value,
        new_value=new_value,
        status="written",
    )


def _pay_plus_group_key(
    resolved: ExcelResolvedValue,
) -> tuple[Path, str, int, str, str, date]:
    planned = resolved.planned_value
    target = planned.target

    return (
        resolved.workbook_path,
        resolved.sheet_name,
        resolved.row_number,
        target.supplier,
        target.location_id,
        target.business_date,
    )

def _mobile_group_key(resolved: ExcelResolvedValue) -> tuple[Path, str, int, str, str, date]:
    planned = resolved.planned_value
    target = planned.target

    return (
        resolved.workbook_path,
        resolved.sheet_name,
        resolved.row_number,
        target.supplier,
        target.location_id,
        target.business_date,
    )


def _apply_mobile_adjustment_group(
    *,
    wb,
    output_path: Path,
    group_values: list[ExcelResolvedValue],
    warnings: list[str],
) -> list[ExcelAppliedChange]:
    changes: list[ExcelAppliedChange] = []

    by_field = {
        resolved.planned_value.field_name: resolved
        for resolved in group_values
    }

    gross_resolved = by_field.get("gross_amt")
    net_resolved = by_field.get("net_amt")
    mobile_resolved = by_field.get("mobile_pay_added_to_gross_net")

    if gross_resolved is None or net_resolved is None or mobile_resolved is None:
        warning = (
            f"Incomplete mobile adjustment group in workbook={output_path.name}: "
            f"fields={sorted(by_field)}"
        )
        warnings.append(warning)
        return changes

    ws = wb[gross_resolved.sheet_name]

    gross_cell = ws[gross_resolved.cell_ref]
    net_cell = ws[net_resolved.cell_ref]
    mobile_cell = ws[mobile_resolved.cell_ref]

    gross_old = gross_cell.value
    net_old = net_cell.value
    mobile_old = mobile_cell.value

    existing_mobile_net = _coerce_decimal(mobile_old)
    planned_mobile_gross = gross_resolved.planned_value.value
    planned_mobile_net = net_resolved.planned_value.value

    target = gross_resolved.planned_value.target

    already_applied = _decimal_equal(existing_mobile_net, planned_mobile_net)
    has_different_existing_mobile = (
        existing_mobile_net is not None
        and not _is_effectively_zero(existing_mobile_net)
        and not already_applied
    )

    if already_applied:
        # Make the mobile column deterministic, but do not add again.
        mobile_cell.value = _to_excel_number(planned_mobile_net)

        for resolved in [gross_resolved, net_resolved, mobile_resolved]:
            planned = resolved.planned_value
            cell = ws[resolved.cell_ref]
            changes.append(
                ExcelAppliedChange(
                    supplier=target.supplier,
                    location_id=target.location_id,
                    location_name=target.location_name,
                    business_date=target.business_date,
                    workbook_path=resolved.workbook_path,
                    output_path=output_path,
                    sheet_name=resolved.sheet_name,
                    cell_ref=resolved.cell_ref,
                    field_name=planned.field_name,
                    column_header=planned.column_header,
                    source=planned.source,
                    mode=planned.mode,
                    old_value=cell.value,
                    new_value=cell.value,
                    status="skipped_already_applied",
                )
            )

        return changes

    if has_different_existing_mobile:
        warning = (
            f"Existing mobile value differs from planned value. "
            f"Skipping mobile add to avoid double-counting/corruption: "
            f"workbook={output_path.name}, sheet={gross_resolved.sheet_name}, "
            f"date={target.business_date}, location={target.location_id}, "
            f"existing_mobile={existing_mobile_net}, planned_mobile_net={planned_mobile_net}"
        )
        warnings.append(warning)

        for resolved in [gross_resolved, net_resolved, mobile_resolved]:
            planned = resolved.planned_value
            cell = ws[resolved.cell_ref]
            changes.append(
                ExcelAppliedChange(
                    supplier=target.supplier,
                    location_id=target.location_id,
                    location_name=target.location_name,
                    business_date=target.business_date,
                    workbook_path=resolved.workbook_path,
                    output_path=output_path,
                    sheet_name=resolved.sheet_name,
                    cell_ref=resolved.cell_ref,
                    field_name=planned.field_name,
                    column_header=planned.column_header,
                    source=planned.source,
                    mode=planned.mode,
                    old_value=cell.value,
                    new_value=cell.value,
                    status="skipped_existing_mobile_mismatch",
                )
            )

        return changes

    gross_existing = _coerce_decimal(gross_old)
    net_existing = _coerce_decimal(net_old)

    if gross_existing is None or net_existing is None:
        warning = (
            f"Cannot apply mobile adjustment because Gross/NET cell is not numeric: "
            f"workbook={output_path.name}, sheet={gross_resolved.sheet_name}, "
            f"date={target.business_date}, location={target.location_id}, "
            f"gross={gross_old!r}, net={net_old!r}"
        )
        warnings.append(warning)
        return changes

    if _is_formula_value(gross_old) or _is_formula_value(net_old):
        warning = (
            f"Cannot apply mobile adjustment over formula Gross/NET cells: "
            f"workbook={output_path.name}, sheet={gross_resolved.sheet_name}, "
            f"date={target.business_date}, location={target.location_id}, "
            f"gross={gross_old!r}, net={net_old!r}"
        )
        warnings.append(warning)
        return changes

    gross_new = _to_excel_number(gross_existing + planned_mobile_gross)
    net_new = _to_excel_number(net_existing + planned_mobile_net)
    mobile_new = _to_excel_number(planned_mobile_net)

    gross_cell.value = gross_new
    net_cell.value = net_new
    mobile_cell.value = mobile_new

    for resolved, old_value, new_value in [
        (gross_resolved, gross_old, gross_new),
        (net_resolved, net_old, net_new),
        (mobile_resolved, mobile_old, mobile_new),
    ]:
        planned = resolved.planned_value
        changes.append(
            ExcelAppliedChange(
                supplier=target.supplier,
                location_id=target.location_id,
                location_name=target.location_name,
                business_date=target.business_date,
                workbook_path=resolved.workbook_path,
                output_path=output_path,
                sheet_name=resolved.sheet_name,
                cell_ref=resolved.cell_ref,
                field_name=planned.field_name,
                column_header=planned.column_header,
                source=planned.source,
                mode=planned.mode,
                old_value=old_value,
                new_value=new_value,
                status="written",
            )
        )

    return changes

def apply_excel_write_plan(
    *,
    plan: ExcelWritePlan,
    resolution: ExcelPlanResolution,
    output_root: Path,
    dry_run: bool,
    write_originals: bool = False,
    backup_originals: bool = True,
) -> ExcelApplyResult:
    apply_result = ExcelApplyResult()

    if resolution.warnings:
        apply_result.warnings.extend(
            "Resolution warning before apply: " + warning
            for warning in resolution.warnings
        )

    if dry_run:
        return apply_result

    resolved_workbooks = {
        resolved.workbook_path
        for resolved in resolution.resolved_values
    }

    output_paths = _prepare_workbook_targets(
        workbook_paths=resolved_workbooks,
        output_root=output_root,
        report_date=plan.report_date,
        write_originals=write_originals,
        backup_originals=backup_originals,
        overwrite_output_copies=True,
    )

    resolved_by_output_workbook: dict[Path, list[ExcelResolvedValue]] = defaultdict(list)

    for resolved in resolution.resolved_values:
        output_path = output_paths[resolved.workbook_path]
        resolved_by_output_workbook[output_path].append(resolved)

    for output_path, workbook_values in resolved_by_output_workbook.items():
        try:
            wb = load_workbook(output_path, data_only=False)
        except Exception as exc:
            apply_result.warnings.append(
                f"Could not open output workbook={output_path}. Error: {exc}"
            )
            continue

        simple_set_values: list[ExcelResolvedValue] = []
        backdated_daily_values: list[ExcelResolvedValue] = []
        mobile_values_by_group = defaultdict(list)
        pay_plus_values_by_group = defaultdict(list)
        monthly_charge_values_by_group = defaultdict(list)

        for resolved in workbook_values:
            planned = resolved.planned_value

            if planned.source == "backdated_daily_total":
                backdated_daily_values.append(resolved)
            elif planned.source == "mobile_adjustment_summary":
                mobile_values_by_group[_mobile_group_key(resolved)].append(resolved)
            elif planned.source == "valero_pay_plus_summary":
                pay_plus_values_by_group[_pay_plus_group_key(resolved)].append(resolved)
            elif planned.source == "valero_monthly_charge_summary":
                monthly_charge_values_by_group[
                    _monthly_charge_group_key(resolved)
                ].append(resolved)
            else:
                simple_set_values.append(resolved)

        for resolved in simple_set_values:
            ws = wb[resolved.sheet_name]
            change = _apply_set_value(
                ws=ws,
                resolved=resolved,
                output_path=output_path,
                warnings=apply_result.warnings,
            )
            apply_result.changes.append(change)

        for resolved in backdated_daily_values:
            change = _apply_backdated_daily_total(
                wb=wb,
                output_path=output_path,
                resolved=resolved,
                warnings=apply_result.warnings,
            )
            apply_result.changes.append(change)

        for group_values in mobile_values_by_group.values():
            changes = _apply_mobile_adjustment_group(
                wb=wb,
                output_path=output_path,
                group_values=group_values,
                warnings=apply_result.warnings,
            )
            apply_result.changes.extend(changes)

        for group_values in monthly_charge_values_by_group.values():
            changes = _apply_valero_monthly_charge_group(
                wb=wb,
                output_path=output_path,
                group_values=group_values,
                warnings=apply_result.warnings,
            )
            apply_result.changes.extend(changes)

        for group_values in pay_plus_values_by_group.values():
            changes = _apply_valero_pay_plus_group(
                wb=wb,
                output_path=output_path,
                group_values=group_values,
                warnings=apply_result.warnings,
            )
            apply_result.changes.extend(changes)

        for group_values in monthly_charge_values_by_group.values():
            changes = _apply_valero_monthly_charge_group(
                wb=wb,
                output_path=output_path,
                group_values=group_values,
                warnings=apply_result.warnings,
            )
            apply_result.changes.extend(changes)

        wb.save(output_path)

    return apply_result


def preview_excel_apply_result(apply_result: ExcelApplyResult) -> None:
    print("\n========== EXCEL APPLY RESULT ==========")
    print(f"written_count={apply_result.written_count}")
    print(f"skipped_count={apply_result.skipped_count}")
    print(f"warnings={len(apply_result.warnings)}")

    print("\n========== APPLIED CELL CHANGES ==========")
    for change in apply_result.changes:
        print(
            f"{change.status}: {change.supplier} | {change.location_id} | "
            f"{change.location_name} | {change.business_date} | "
            f"{change.output_path.name} | {change.sheet_name}!{change.cell_ref} | "
            f"{change.column_header} | {change.source} | {change.mode} | "
            f"{change.old_value!r} -> {change.new_value!r}"
        )

    if apply_result.warnings:
        print("\n========== EXCEL APPLY WARNINGS ==========")
        for warning in apply_result.warnings:
            print(f"WARNING: {warning}")

def _get_backup_workbook_path(
    *,
    original_workbook_path: Path,
    output_root: Path,
    report_date: date,
) -> Path:
    return (
        output_root
        / "_backups"
        / report_date.isoformat()
        / original_workbook_path.name
    )


def _prepare_workbook_targets(
    *,
    workbook_paths: set[Path],
    output_root: Path,
    report_date: date,
    write_originals: bool,
    backup_originals: bool,
    overwrite_output_copies: bool = True,
) -> dict[Path, Path]:
    """
    Returns:
        original workbook path -> workbook path that should be edited

    Copy mode:
        data/workbooks/A.xlsx -> output/excel/<date>/A.xlsx

    Original mode:
        data/workbooks/A.xlsx -> data/workbooks/A.xlsx
        optionally creates backup under output/excel/_backups/<date>/A.xlsx
    """
    target_paths: dict[Path, Path] = {}

    for workbook_path in workbook_paths:
        if write_originals:
            if backup_originals:
                backup_path = _get_backup_workbook_path(
                    original_workbook_path=workbook_path,
                    output_root=output_root,
                    report_date=report_date,
                )
                backup_path.parent.mkdir(parents=True, exist_ok=True)

                if not backup_path.exists():
                    copy2(workbook_path, backup_path)

            target_paths[workbook_path] = workbook_path
            continue

        output_path = _get_output_workbook_path(
            original_workbook_path=workbook_path,
            output_root=output_root,
            report_date=report_date,
        )
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if overwrite_output_copies or not output_path.exists():
            copy2(workbook_path, output_path)

        target_paths[workbook_path] = output_path

    return target_paths

def build_excel_write_plan(
    *,
    report: ParsedReport,
    workbook_root: Path,
) -> ExcelWritePlan:
    supplier = normalize_supplier(report.supplier)

    plan = ExcelWritePlan(
        report_supplier=supplier,
        report_date=report.report_date,
    )

    primary_daily_dates = _get_primary_daily_total_dates(report)

    for row in report.daily_totals:
        if row.date in primary_daily_dates:
            _append_daily_total_values(
                plan=plan,
                workbook_root=workbook_root,
                row=row,
                mode="set",
                source="daily_total",
            )
        else:
            _append_daily_total_values(
                plan=plan,
                workbook_root=workbook_root,
                row=row,
                mode="add_to_base",
                source="backdated_daily_total",
            )

    if supplier == "VALERO":
        mobile_summary_rows = summarize_mobile_adjustments(report.mobile_adjustments)

        for row in mobile_summary_rows:
            _append_valero_mobile_summary_values(
                plan=plan,
                workbook_root=workbook_root,
                row=row,
            )

        valero_pay_plus_rows = getattr(report, "valero_pay_plus_adjustments", [])
        valero_pay_plus_summary_rows = summarize_valero_pay_plus_adjustments(
            valero_pay_plus_rows
        )

        for row in valero_pay_plus_summary_rows:
            _append_valero_pay_plus_values(
                plan=plan,
                workbook_root=workbook_root,
                row=row,
            )

        monthly_charge_rows = getattr(report, "valero_monthly_charges", [])
        monthly_charge_summary_rows = _summarize_valero_monthly_charges(
            monthly_charge_rows
        )

        for row in monthly_charge_summary_rows:
            _append_valero_monthly_charge_values(
                plan=plan,
                workbook_root=workbook_root,
                row=row,
            )
    else:
        if report.mobile_adjustments:
            plan.warnings.append(
                f"Ignoring {len(report.mobile_adjustments)} mobile adjustment rows "
                f"for non-Valero supplier={supplier}."
            )

        valero_pay_plus_rows = getattr(report, "valero_pay_plus_adjustments", [])

        if valero_pay_plus_rows:
            plan.warnings.append(
                f"Ignoring {len(valero_pay_plus_rows)} Valero Pay+ rows "
                f"for non-Valero supplier={supplier}."
            )

        monthly_charge_rows = getattr(report, "valero_monthly_charges", [])

        if monthly_charge_rows:
            plan.warnings.append(
                f"Ignoring {len(monthly_charge_rows)} Valero monthly charge rows "
                f"for non-Valero supplier={supplier}."
            )

    return plan


def preview_excel_write_plan(plan: ExcelWritePlan) -> None:
    print("\n========== EXCEL WRITE PLAN ==========")
    print(f"report_supplier={plan.report_supplier}")
    print(f"report_date={plan.report_date}")
    print(f"planned_values={len(plan.planned_values)}")
    print(f"fee_validations={len(plan.fee_validations)}")
    print(f"affected_workbooks={len(plan.affected_workbooks)}")

    for workbook_path in sorted(plan.affected_workbooks):
        print(f"  workbook={workbook_path}")

    print("\n========== PLANNED CELL VALUES ==========")
    for planned in plan.planned_values:
        target = planned.target
        print(
            f"{target.supplier} | {target.location_id} | {target.location_name} | "
            f"{target.business_date} | {planned.column_header} | "
            f"{planned.source} | {planned.mode} | {planned.value} | "
            f"{target.workbook_path.name}"
        )

    print("\n========== FEE VALIDATION SUMMARY ==========")
    failed_fee_validations = [
        validation for validation in plan.fee_validations if not validation.is_valid
    ]
    print(f"total_fee_validations={len(plan.fee_validations)}")
    print(f"failed_fee_validations={len(failed_fee_validations)}")

    for validation in failed_fee_validations:
        print(
            "FEE WARNING: "
            f"{validation.supplier} | {validation.location_id} | "
            f"{validation.location_name} | {validation.business_date} | "
            f"{validation.source} | "
            f"gross={validation.gross_amt} | "
            f"net={validation.net_amt} | "
            f"calculated_fee={validation.calculated_fees} | "
            f"parsed_fee={validation.parsed_fees} | "
            f"difference={validation.difference}"
        )

    if plan.warnings:
        print("\n========== EXCEL PLAN WARNINGS ==========")
        for warning in plan.warnings:
            print(f"WARNING: {warning}")


def write_parsed_report_to_excel(
    *,
    report: ParsedReport,
    workbook_root: Path,
    output_root: Path,
    dry_run: bool = True,
    write_originals: bool = False,
    backup_originals: bool = True,
) -> ExcelWriteResult:
    plan = build_excel_write_plan(
        report=report,
        workbook_root=workbook_root,
    )

    preview_excel_write_plan(plan)

    resolution = resolve_excel_write_plan_targets(plan)
    preview_excel_resolution(resolution)

    if resolution.warnings:
        print(
            "\n[WARNING] Resolution produced warnings. "
            "Continuing only if dry_run=True; write mode will still copy files "
            "but may skip unsafe cells."
        )

    apply_result = apply_excel_write_plan(
        plan=plan,
        resolution=resolution,
        output_root=output_root,
        dry_run=dry_run,
        write_originals=write_originals,
        backup_originals=backup_originals,
    )
    preview_excel_apply_result(apply_result)

    return ExcelWriteResult(
        dry_run=dry_run,
        write_originals=write_originals,
        plan=plan,
        resolution=resolution,
        apply_result=apply_result,
        written_count=apply_result.written_count,
        skipped_count=apply_result.skipped_count,
        warnings=(
                plan.warnings.copy()
                + resolution.warnings.copy()
                + apply_result.warnings.copy()
        ),
    )


def _validate_fee_math(
    *,
    plan: ExcelWritePlan,
    supplier: str,
    location_id: str,
    location_name: str,
    business_date: date,
    gross_amt: Decimal,
    net_amt: Decimal,
    parsed_fees: Decimal,
    source: str,
) -> None:
    policy = EXCEL_MAPPING.fee_validation_policy

    if not policy.enabled:
        return

    calculated_fees = gross_amt - net_amt
    difference = calculated_fees - parsed_fees
    is_valid = abs(difference) <= policy.tolerance

    validation = ExcelFeeValidation(
        supplier=normalize_supplier(supplier),
        location_id=str(location_id),
        location_name=location_name,
        business_date=business_date,
        source=source,
        gross_amt=gross_amt,
        net_amt=net_amt,
        parsed_fees=parsed_fees,
        calculated_fees=calculated_fees,
        difference=difference,
        is_valid=is_valid,
    )

    plan.fee_validations.append(validation)

    if not is_valid:
        plan.warnings.append(
            "Fee validation mismatch: "
            f"{validation.supplier} {validation.location_id} "
            f"{validation.location_name} date={validation.business_date} "
            f"source={validation.source} "
            f"gross={validation.gross_amt} net={validation.net_amt} "
            f"calculated_fee={validation.calculated_fees} "
            f"parsed_fee={validation.parsed_fees} "
            f"difference={validation.difference}"
        )


def resolve_excel_write_plan_targets(plan: ExcelWritePlan) -> ExcelPlanResolution:
    resolution = ExcelPlanResolution()

    values_by_workbook: dict[Path, list[ExcelPlannedValue]] = defaultdict(list)

    for planned_value in plan.planned_values:
        values_by_workbook[planned_value.target.workbook_path].append(planned_value)

    for expected_workbook_path, planned_values in values_by_workbook.items():
        workbook_path, workbook_warnings = _resolve_existing_workbook_path(
            expected_workbook_path
        )
        resolution.warnings.extend(workbook_warnings)

        if workbook_path is None:
            continue

        try:
            # Formula workbook: used for headers, formulas, current values, and writing later.
            wb = load_workbook(workbook_path, data_only=False)

            # Value workbook: used only for resolving date rows whose cells may be formulas.
            wb_values = load_workbook(workbook_path, data_only=True)
        except Exception as exc:
            resolution.warnings.append(
                f"Could not open workbook: {workbook_path}. Error: {exc}"
            )
            continue

        values_by_date: dict[date, list[ExcelPlannedValue]] = defaultdict(list)
        for planned_value in planned_values:
            values_by_date[planned_value.target.business_date].append(planned_value)

        for business_date, date_values in values_by_date.items():
            sheet_name = resolve_month_sheet_name(wb.sheetnames, business_date)

            if sheet_name is None:
                resolution.warnings.append(
                    f"Missing month sheet for date={business_date} "
                    f"in workbook={workbook_path.name}. "
                    f"Available sheets={wb.sheetnames}"
                )
                continue

            ws = wb[sheet_name]
            ws_values = wb_values[sheet_name]

            required_headers = {
                EXCEL_MAPPING.columns.date,
                EXCEL_MAPPING.columns.gross_amt,
                EXCEL_MAPPING.columns.net_amt,
                EXCEL_MAPPING.columns.fees,
            }

            for planned_value in date_values:
                required_headers.add(planned_value.column_header)

            header_result = _find_header_row_and_columns(
                ws,
                required_headers,
                max_rows=EXCEL_MAPPING.header_scan_max_rows,
            )

            if header_result is None:
                resolution.warnings.append(
                    f"Could not find required headers in workbook={workbook_path.name}, "
                    f"sheet={sheet_name}, required={sorted(required_headers)}"
                )
                continue

            header_row, header_map = header_result
            date_column_number = header_map[normalize_excel_text(EXCEL_MAPPING.columns.date)]

            row_number = _find_date_row(
                ws_values,
                date_column_number=date_column_number,
                header_row=header_row,
                target_date=business_date,
            )

            if row_number is None:
                resolution.warnings.append(
                    f"Could not find date row: workbook={workbook_path.name}, "
                    f"sheet={sheet_name}, date={business_date}"
                )
                continue

            fee_column_number = header_map[normalize_excel_text(EXCEL_MAPPING.columns.fees)]
            fee_cell = ws.cell(row=row_number, column=fee_column_number)

            first_value = date_values[0]
            fee_validation = ExcelFeeFormulaValidation(
                supplier=first_value.target.supplier,
                location_id=first_value.target.location_id,
                location_name=first_value.target.location_name,
                business_date=business_date,
                workbook_path=workbook_path,
                sheet_name=sheet_name,
                cell_ref=fee_cell.coordinate,
                current_value=fee_cell.value,
                is_formula=_is_formula_value(fee_cell.value),
            )
            resolution.fee_formula_validations.append(fee_validation)

            if (
                EXCEL_MAPPING.fee_validation_policy.require_formula_cell
                and not fee_validation.is_formula
            ):
                resolution.warnings.append(
                    f"CC Fee cell is not a formula: workbook={workbook_path.name}, "
                    f"sheet={sheet_name}, date={business_date}, "
                    f"cell={fee_cell.coordinate}, value={fee_cell.value!r}"
                )

            for planned_value in date_values:
                normalized_header = normalize_excel_text(planned_value.column_header)

                if normalized_header not in header_map:
                    resolution.warnings.append(
                        f"Missing column header={planned_value.column_header} "
                        f"in workbook={workbook_path.name}, sheet={sheet_name}"
                    )
                    continue

                column_number = header_map[normalized_header]
                cell = ws.cell(row=row_number, column=column_number)
                cell_ref = f"{get_column_letter(column_number)}{row_number}"

                resolution.resolved_values.append(
                    ExcelResolvedValue(
                        planned_value=planned_value,
                        workbook_path=workbook_path,
                        sheet_name=sheet_name,
                        header_row=header_row,
                        row_number=row_number,
                        column_number=column_number,
                        cell_ref=cell_ref,
                        current_value=cell.value,
                    )
                )

    return resolution

def preview_excel_resolution(resolution: ExcelPlanResolution) -> None:
    print("\n========== EXCEL TARGET RESOLUTION ==========")
    print(f"resolved_values={len(resolution.resolved_values)}")
    print(f"fee_formula_validations={len(resolution.fee_formula_validations)}")
    print(f"warnings={len(resolution.warnings)}")

    print("\n========== RESOLVED CELL TARGETS ==========")
    for resolved in resolution.resolved_values:
        planned = resolved.planned_value
        target = planned.target
        print(
            f"{target.supplier} | {target.location_id} | {target.location_name} | "
            f"{target.business_date} | {resolved.workbook_path.name} | "
            f"{resolved.sheet_name}!{resolved.cell_ref} | "
            f"{planned.column_header} | {planned.source} | {planned.mode} | "
            f"current={resolved.current_value!r} | planned={planned.value}"
        )

    print("\n========== CC FEE FORMULA VALIDATION ==========")
    for validation in resolution.fee_formula_validations:
        status = "OK" if validation.is_formula else "WARNING"
        print(
            f"{status}: {validation.supplier} | {validation.location_id} | "
            f"{validation.location_name} | {validation.business_date} | "
            f"{validation.workbook_path.name} | "
            f"{validation.sheet_name}!{validation.cell_ref} | "
            f"value={validation.current_value!r}"
        )

    if resolution.warnings:
        print("\n========== EXCEL RESOLUTION WARNINGS ==========")
        for warning in resolution.warnings:
            print(f"WARNING: {warning}")