from __future__ import annotations
from _path_setup import PROJECT_ROOT  # noqa: F401
import argparse
import calendar
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from dataclasses import dataclass
from typing import Mapping



from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from config.locations import (
    CITGO_LOCATIONS,
    VALERO_LOCATIONS,
    SUNOCO_DEALER_LOCATIONS,
    SUNOCO_WHOLESALER_LOCATIONS,
)

MONTH_ABBR = [
    "JAN",
    "FEB",
    "MAR",
    "APR",
    "MAY",
    "JUN",
    "JUL",
    "AUG",
    "SEP",
    "OCT",
    "NOV",
    "DEC",
]


from dataclasses import dataclass
from typing import Mapping

from config.locations import CITGO_LOCATIONS, VALERO_LOCATIONS, SUNOCO_LOCATIONS


VALERO_COLUMNS = [
    "DATE",
    "STORE AMT",
    "DIFF",
    "GROSS AMT",
    "NET AMT",
    "CC FEE",
    "DATE OF VALERO CREDIT",
    "MONTHLY VAL CHGS IN FEES",
    "FUELMAN",
    "MOBILE PAY ADDED TO GROSS/NET",
    "VALERO PAY +",
    "GIFT CARDS",
]


SUNOCO_DEALER_COLUMNS = [
    "DATE",
    "STORE AMT",
    "DIFF",
    "GROSS AMT",
    "NET AMT",
    "CC FEE",
    "DATE OF SUNOCO CREDIT",
    "SUNOCO MONTHLY CHGS",
    "GIFT CARD",
]


SUNOCO_WHOLESALER_COLUMNS = [
    "DATE",
    "GROSS AMT",
    "NET AMT",
    "DATE OF SUNOCO CREDIT",
    "CC FEE",
    "MONTHLY BILLING",
    "GIFT CARD DEDUCTED FROM GROSS/NET",
    "DATE INV APPLIED",
]


CITGO_COLUMNS = [
    "DATE",
    "STORE AMT",
    "DIFF",
    "GROSS AMT",
    "NET AMT",
    "CC FEE",
    "DATE OF VALERO CREDIT",
    "MONTHLY CITGO CHGS IN FEES",
    "MOBILE PAY",
    "MOBILE PAY DISC",
    "FUEL DISC",
    "CHARGEBACK DED ON DRAFT",
    "LOYALTY FEE",
]



@dataclass(frozen=True)
class SupplierWorkbookConfig:
    supplier: str
    workbook_type: str
    locations: Mapping[str, str]
    columns: list[str]

    # Optional formula behavior.
    diff_formula: str | None = None
    cc_fee_formula: str | None = None

    # Columns that should be treated as dates and should not be summed.
    date_columns: tuple[str, ...] = ("DATE",)


SUPPLIER_WORKBOOK_CONFIGS: dict[str, SupplierWorkbookConfig] = {
    "CITGO": SupplierWorkbookConfig(
        supplier="CITGO",
        workbook_type="CITGO",
        locations=CITGO_LOCATIONS,
        columns=CITGO_COLUMNS,
        diff_formula="gross_minus_store",
        cc_fee_formula="gross_minus_net",
        date_columns=("DATE", "DATE OF VALERO CREDIT"),
    ),

    "VALERO": SupplierWorkbookConfig(
        supplier="VALERO",
        workbook_type="VALERO",
        locations=VALERO_LOCATIONS,
        columns=VALERO_COLUMNS,
        diff_formula="store_minus_gross",
        cc_fee_formula="gross_minus_net",
        date_columns=("DATE", "DATE OF VALERO CREDIT"),
    ),

    "SUNOCO_DEALER": SupplierWorkbookConfig(
        supplier="SUNOCO",
        workbook_type="SUNOCO_DEALER",
        locations=SUNOCO_DEALER_LOCATIONS,
        columns=SUNOCO_DEALER_COLUMNS,
        diff_formula="gross_minus_store",
        cc_fee_formula="net_minus_gross",
        date_columns=("DATE", "DATE OF SUNOCO CREDIT"),
    ),

    "SUNOCO_WHOLESALER": SupplierWorkbookConfig(
        supplier="SUNOCO",
        workbook_type="SUNOCO_WHOLESALER",
        locations=SUNOCO_WHOLESALER_LOCATIONS,
        columns=SUNOCO_WHOLESALER_COLUMNS,
        diff_formula=None,
        cc_fee_formula="net_minus_gross",
        date_columns=("DATE", "DATE OF SUNOCO CREDIT", "DATE INV APPLIED"),
    ),
}


@dataclass(frozen=True)
class SettlementWorkbookTarget:
    supplier: str
    workbook_type: str
    location_id: str
    location_name: str
    columns: list[str]
    diff_formula: str | None
    cc_fee_formula: str | None
    date_columns: tuple[str, ...]

WORKBOOK_TARGETS: list[SettlementWorkbookTarget] = [
    SettlementWorkbookTarget(
        supplier=config.supplier,
        workbook_type=config.workbook_type,
        location_id=location_id,
        location_name=location_name,
        columns=config.columns,
        diff_formula=config.diff_formula,
        cc_fee_formula=config.cc_fee_formula,
        date_columns=config.date_columns,
    )
    for config in SUPPLIER_WORKBOOK_CONFIGS.values()
    for location_id, location_name in config.locations.items()
]
# Keep this mapping aligned with config/excel_mapping.py and config/locations.py.
# The filename uses location_name, not location_id.

LOCATION_MAPS = {
    "CITGO": CITGO_LOCATIONS,
    "VALERO": VALERO_LOCATIONS,
    "SUNOCO": SUNOCO_LOCATIONS,
}



def normalize_for_filename(value: str) -> str:
    """
    Normalize names for workbook filenames.

    Keeps spaces because the existing naming convention uses spaces:
        2026 CC MONTECELLO VALERO.xlsx
    """
    value = value.strip().upper()
    value = re.sub(r"[\\/:*?\"<>|]", "", value)
    value = re.sub(r"\s+", " ", value)
    return value


def build_workbook_filename(year: int, location_name: str, supplier: str) -> str:
    location_name = normalize_for_filename(location_name)
    supplier = normalize_for_filename(supplier)

    if location_name.endswith(supplier):
        return f"{year} CC {location_name}.xlsx"

    return f"{year} CC {location_name} {supplier}.xlsx"
def build_sheet_name(year: int, month: int) -> str:
    return f"{MONTH_ABBR[month - 1]} {year}"


def style_month_sheet(
    ws,
    year: int,
    month: int,
    supplier: str,
    workbook_type: str,
    location_name: str,
    columns: list[str],
    diff_formula: str | None,
    cc_fee_formula: str | None,
    date_columns: tuple[str, ...],
) -> None:
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_col = len(columns)
    last_col_letter = get_column_letter(max_col)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"] = (
        f"{supplier.upper()} SETTLEMENT - "
        f"{location_name.upper()} - "
        f"{build_sheet_name(year, month)}"
    )
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def col_index(header: str) -> int | None:
        if header not in columns:
            return None
        return columns.index(header) + 1

    date_col = col_index("DATE")
    store_amt_col = col_index("STORE AMT")
    diff_col = col_index("DIFF")
    gross_col = col_index("GROSS AMT")
    net_col = col_index("NET AMT")
    cc_fee_col = col_index("CC FEE")

    days_in_month = calendar.monthrange(year, month)[1]

    for day in range(1, days_in_month + 1):
        row = day + 2
        current_date = date(year, month, day)

        if date_col is not None:
            ws.cell(row=row, column=date_col).value = current_date
            ws.cell(row=row, column=date_col).number_format = "m/d/yyyy"

        # DIFF formula.
        # Dealer SUNOCO:
        #   DIFF = GROSS AMT - STORE AMT
        #
        # Valero:
        #   DIFF = STORE AMT - GROSS AMT
        #
        # Wholesaler SUNOCO:
        #   no DIFF column, so this section is skipped.
        if (
            diff_formula is not None
            and diff_col is not None
            and store_amt_col is not None
            and gross_col is not None
        ):
            store_letter = get_column_letter(store_amt_col)
            gross_letter = get_column_letter(gross_col)

            if diff_formula == "gross_minus_store":
                diff_expr = f"{gross_letter}{row}-{store_letter}{row}"
            elif diff_formula == "store_minus_gross":
                diff_expr = f"{store_letter}{row}-{gross_letter}{row}"
            else:
                raise ValueError(f"Unsupported diff formula: {diff_formula}")

            ws.cell(row=row, column=diff_col).value = (
                f'=IF(OR({store_letter}{row}="",{gross_letter}{row}=""),"",'
                f'ROUND({diff_expr},2))'
            )

        # CC FEE formula.
        # SUNOCO:
        #   CC FEE = NET AMT - GROSS AMT
        #
        # Valero/CITGO:
        #   CC FEE = GROSS AMT - NET AMT
        if (
            cc_fee_formula is not None
            and cc_fee_col is not None
            and gross_col is not None
            and net_col is not None
        ):
            gross_letter = get_column_letter(gross_col)
            net_letter = get_column_letter(net_col)

            if cc_fee_formula == "net_minus_gross":
                cc_fee_expr = f"{net_letter}{row}-{gross_letter}{row}"
            elif cc_fee_formula == "gross_minus_net":
                cc_fee_expr = f"{gross_letter}{row}-{net_letter}{row}"
            else:
                raise ValueError(f"Unsupported CC fee formula: {cc_fee_formula}")

            ws.cell(row=row, column=cc_fee_col).value = (
                f'=IF(OR({gross_letter}{row}="",{net_letter}{row}=""),"",'
                f'ROUND({cc_fee_expr},2))'
            )

        for col_idx, header in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if header in date_columns:
                cell.number_format = "m/d/yyyy"
            else:
                cell.number_format = '#,##0.00'

    total_row = days_in_month + 4
    ws.cell(row=total_row, column=1).value = "MONTH TOTAL"
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = border

        if header in date_columns:
            continue

        col_letter = get_column_letter(col_idx)
        cell.value = f"=SUM({col_letter}3:{col_letter}{days_in_month + 2})"
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last_col_letter}{days_in_month + 2}"

    for col_idx, header in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)

        if header in date_columns:
            width = 18
        elif len(header) >= 28:
            width = 34
        elif len(header) >= 20:
            width = 26
        else:
            width = 14

        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 42


def create_settlement_workbook(
    output_path: Path,
    year: int,
    supplier: str,
    workbook_type: str,
    location_name: str,
    columns: list[str],
    diff_formula: str | None,
    cc_fee_formula: str | None,
    date_columns: tuple[str, ...],
) -> None:
    wb = Workbook()

    default_ws = wb.active
    wb.remove(default_ws)

    for month in range(1, 13):
        ws = wb.create_sheet(title=build_sheet_name(year, month))
        style_month_sheet(
            ws=ws,
            year=year,
            month=month,
            supplier=supplier,
            workbook_type=workbook_type,
            location_name=location_name,
            columns=columns,
            diff_formula=diff_formula,
            cc_fee_formula=cc_fee_formula,
            date_columns=date_columns,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)



def filter_targets(
    supplier: str | None,
    location_id: str | None,
    location_name: str | None,
) -> list[SettlementWorkbookTarget]:
    targets = WORKBOOK_TARGETS

    if supplier:
        supplier_upper = supplier.upper()
        targets = [t for t in targets if t.supplier.upper() == supplier_upper]

    if location_id:
        targets = [t for t in targets if t.location_id == location_id]

    if location_name:
        name_upper = location_name.upper()
        targets = [t for t in targets if t.location_name.upper() == name_upper]

    return targets


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Create blank settlement Excel workbooks."
    )
    parser.add_argument(
        "--year",
        type=int,
        required=True,
        help="Settlement workbook year, e.g. 2026.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data/excel/settlements"),
        help="Directory where workbooks should be created.",
    )
    parser.add_argument(
        "--supplier",
        help="Optional supplier filter, e.g. VALERO, CITGO, SUNOCO.",
    )
    parser.add_argument(
        "--location-id",
        help="Optional location id filter, e.g. 15861002.",
    )
    parser.add_argument(
        "--location-name",
        help="Optional location name filter, e.g. HAVERSTRAW.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing workbooks.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print files that would be created without creating them.",
    )

    args = parser.parse_args()

    targets = filter_targets(
        supplier=args.supplier,
        location_id=args.location_id,
        location_name=args.location_name,
    )

    if not targets:
        raise SystemExit("No workbook targets matched the given filters.")

    created = 0
    skipped = 0

    for target in targets:
        filename = build_workbook_filename(
            year=args.year,
            location_name=target.location_name,
            supplier=target.supplier,
        )
        output_path = args.output_dir / filename

        if args.dry_run:
            print(f"[DRY RUN] Would create: {output_path}")
            continue

        if output_path.exists() and not args.overwrite:
            print(f"[SKIP] Exists already: {output_path}")
            skipped += 1
            continue

        create_settlement_workbook(
            output_path=output_path,
            year=args.year,
            supplier=target.supplier,
            workbook_type=target.workbook_type,
            location_name=target.location_name,
            columns=target.columns,
            diff_formula=target.diff_formula,
            cc_fee_formula=target.cc_fee_formula,
            date_columns=target.date_columns,
        )
        print(f"[CREATE] {output_path}")
        created += 1

    print()
    print(f"Created : {created}")
    print(f"Skipped : {skipped}")
    print(f"Dry run : {args.dry_run}")


if __name__ == "__main__":
    main()