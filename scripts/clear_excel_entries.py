from _path_setup import PROJECT_ROOT  # noqa: F401

import argparse
import calendar
import re
from datetime import date, datetime
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.datetime import from_excel


DATE_FORMULA_RE = re.compile(
    r"^\s*=\s*\$?([A-Z]+)\$?(\d+)\s*([+-])\s*(\d+)\s*$",
    re.IGNORECASE,
)


def parse_iso_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise argparse.ArgumentTypeError(
            f"Invalid date: {value}. Expected YYYY-MM-DD."
        ) from exc


def parse_month(value: str) -> tuple[date, date]:
    try:
        year_text, month_text = value.split("-", 1)
        year = int(year_text)
        month = int(month_text)
        last_day = calendar.monthrange(year, month)[1]
        return date(year, month, 1), date(year, month, last_day)
    except Exception as exc:
        raise argparse.ArgumentTypeError(
            f"Invalid month: {value}. Expected YYYY-MM."
        ) from exc


def normalize_excel_text(value: object) -> str:
    if value is None:
        return ""

    text = str(value).strip().upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def is_excel_temp_file(path: Path) -> bool:
    return path.name.startswith("~$")


def coerce_excel_date(value: object) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, int | float):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    if isinstance(value, str):
        text = value.strip()

        if text.startswith("="):
            return None

        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%d-%B-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue

    return None


def evaluate_date_cell(
    ws,
    *,
    row_idx: int,
    column_idx: int,
    header_row: int,
    cache: dict[tuple[int, int], date],
    visiting: set[tuple[int, int]],
) -> date | None:
    key = (row_idx, column_idx)

    if key in cache:
        return cache[key]

    if key in visiting:
        return None

    visiting.add(key)

    cell_value = ws.cell(row=row_idx, column=column_idx).value
    direct_date = coerce_excel_date(cell_value)

    if direct_date is not None:
        cache[key] = direct_date
        visiting.remove(key)
        return direct_date

    if not isinstance(cell_value, str):
        visiting.remove(key)
        return None

    match = DATE_FORMULA_RE.match(cell_value)

    if not match:
        visiting.remove(key)
        return None

    ref_col_letters, ref_row_text, operator, offset_text = match.groups()
    ref_col_idx = column_index_from_string(ref_col_letters.upper())
    ref_row_idx = int(ref_row_text)
    offset_days = int(offset_text)

    if ref_row_idx <= header_row:
        visiting.remove(key)
        return None

    ref_date = evaluate_date_cell(
        ws,
        row_idx=ref_row_idx,
        column_idx=ref_col_idx,
        header_row=header_row,
        cache=cache,
        visiting=visiting,
    )

    if ref_date is None:
        visiting.remove(key)
        return None

    result = (
        ref_date.replace()
        if operator == "+"
        else ref_date.replace()
    )

    if operator == "+":
        from datetime import timedelta

        result = ref_date + timedelta(days=offset_days)
    else:
        from datetime import timedelta

        result = ref_date - timedelta(days=offset_days)

    cache[key] = result
    visiting.remove(key)
    return result


def find_header_row_and_date_column(ws, *, max_rows: int = 15) -> tuple[int, int] | None:
    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        for cell in ws[row_idx]:
            if normalize_excel_text(cell.value) == "DATE":
                return row_idx, cell.column

    return None


def get_header_by_column(ws, header_row: int) -> dict[int, str]:
    headers: dict[int, str] = {}

    for cell in ws[header_row]:
        normalized = normalize_excel_text(cell.value)

        if normalized:
            headers[cell.column] = normalized

    return headers


def clear_workbook(
    *,
    workbook_path: Path,
    output_path: Path,
    start_date: date,
    end_date: date,
    write: bool,
    preserve_formula_headers: set[str],
) -> tuple[int, list[str]]:
    warnings: list[str] = []
    cleared_cells = 0

    try:
        wb = load_workbook(workbook_path, data_only=False)
    except Exception as exc:
        return 0, [f"Could not open workbook={workbook_path.name}: {exc}"]

    for ws in wb.worksheets:
        header_info = find_header_row_and_date_column(ws)

        if header_info is None:
            continue

        header_row, date_column = header_info
        headers_by_column = get_header_by_column(ws, header_row)
        date_cache: dict[tuple[int, int], date] = {}

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_date = evaluate_date_cell(
                ws,
                row_idx=row_idx,
                column_idx=date_column,
                header_row=header_row,
                cache=date_cache,
                visiting=set(),
            )

            if row_date is None:
                continue

            if not (start_date <= row_date <= end_date):
                continue

            for col_idx in range(1, ws.max_column + 1):
                if col_idx == date_column:
                    continue

                header_name = headers_by_column.get(col_idx, "")
                cell = ws.cell(row=row_idx, column=col_idx)

                if (
                    header_name in preserve_formula_headers
                    and isinstance(cell.value, str)
                    and cell.value.strip().startswith("=")
                ):
                    continue

                if cell.value is not None:
                    cell.value = None
                    cleared_cells += 1

    if write:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    return cleared_cells, warnings


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Clear Excel-entered values for a date range while keeping Date column/header structure."
    )

    parser.add_argument(
        "--workbook-root",
        type=Path,
        required=True,
        help="Folder containing Excel workbooks.",
    )

    parser.add_argument(
        "--output-root",
        type=Path,
        default=Path("output/excel_cleared"),
        help="Where cleared workbook copies/backups are written.",
    )

    parser.add_argument(
        "--month",
        type=parse_month,
        help="Month to clear, format YYYY-MM. Example: 2026-07.",
    )

    parser.add_argument(
        "--start-date",
        type=parse_iso_date,
        help="Start date inclusive, format YYYY-MM-DD.",
    )

    parser.add_argument(
        "--end-date",
        type=parse_iso_date,
        help="End date inclusive, format YYYY-MM-DD.",
    )

    parser.add_argument(
        "--write",
        action="store_true",
        help="Actually save cleared workbooks. Without this, dry-run only.",
    )

    parser.add_argument(
        "--write-originals",
        action="store_true",
        help="Modify workbooks in workbook-root directly. Requires --write.",
    )

    parser.add_argument(
        "--no-backup-originals",
        action="store_true",
        help="Do not create backups when using --write-originals.",
    )

    parser.add_argument(
        "--preserve-formula-columns",
        default="CC Fee,DIFF",
        help=(
            "Comma-separated headers whose formulas should be preserved. "
            "Default: 'CC Fee,DIFF'."
        ),
    )

    return parser


def main() -> int:
    args = build_parser().parse_args()

    if args.write_originals and not args.write:
        print("[FAILED] --write-originals requires --write.")
        return 1

    if args.month:
        start_date, end_date = args.month
    else:
        if not args.start_date or not args.end_date:
            print("[FAILED] Provide either --month or both --start-date and --end-date.")
            return 1

        start_date = args.start_date
        end_date = args.end_date

    if start_date > end_date:
        print("[FAILED] start date cannot be after end date.")
        return 1

    preserve_formula_headers = {
        normalize_excel_text(header)
        for header in args.preserve_formula_columns.split(",")
        if header.strip()
    }

    workbooks = sorted(
        path
        for path in args.workbook_root.glob("*.xlsx")
        if not is_excel_temp_file(path)
    )

    print("\nEXCEL ENTRY CLEARER")
    print("=" * 100)
    print(f"workbook_root              : {args.workbook_root}")
    print(f"output_root                : {args.output_root}")
    print(f"start_date                 : {start_date}")
    print(f"end_date                   : {end_date}")
    print(f"dry_run                    : {not args.write}")
    print(f"write_originals            : {args.write_originals}")
    print(f"backup_originals           : {not args.no_backup_originals}")
    print(f"preserve_formula_columns   : {sorted(preserve_formula_headers)}")
    print(f"workbooks_found            : {len(workbooks)}")
    print("=" * 100)

    total_cleared = 0
    total_warnings = 0
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    for workbook_path in workbooks:
        if args.write_originals:
            output_path = workbook_path

            if args.write and not args.no_backup_originals:
                backup_path = (
                    args.output_root
                    / "_clear_backups"
                    / timestamp
                    / workbook_path.name
                )
                backup_path.parent.mkdir(parents=True, exist_ok=True)
                copy2(workbook_path, backup_path)
        else:
            output_path = args.output_root / workbook_path.name

        cleared_cells, warnings = clear_workbook(
            workbook_path=workbook_path,
            output_path=output_path,
            start_date=start_date,
            end_date=end_date,
            write=args.write,
            preserve_formula_headers=preserve_formula_headers,
        )

        total_cleared += cleared_cells
        total_warnings += len(warnings)

        if cleared_cells or warnings:
            print(f"\n{workbook_path.name}")
            print(f"  cleared_cells : {cleared_cells}")
            print(f"  output_path   : {output_path if args.write else '[dry-run]'}")

            for warning in warnings:
                print(f"  WARNING: {warning}")

    print("\nSUMMARY")
    print("=" * 100)
    print(f"workbooks_scanned : {len(workbooks)}")
    print(f"total_cleared     : {total_cleared}")
    print(f"total_warnings    : {total_warnings}")
    print("=" * 100)

    return 0 if total_warnings == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())